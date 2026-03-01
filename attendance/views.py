from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from django.utils import timezone
from django.utils.timezone import localtime
from datetime import timedelta
from .models import Attendance, LeaveRequest
from .serializers import AttendanceSerializer, LeaveRequestSerializer
from employees.models import Employee
import openpyxl
from django.http import HttpResponse


def hitung_hari_leave(leaves, month, year):
    total_izin = 0
    total_sakit = 0
    for leave in leaves:
        start = leave.start_date
        end = leave.end_date
        current = start
        while current <= end:
            if (not month or current.month == int(month)) and \
               (not year or current.year == int(year)) and \
               current.weekday() < 5:
                if leave.type == 'izin':
                    total_izin += 1
                else:
                    total_sakit += 1
            current += timedelta(days=1)
    return total_izin, total_sakit


class AttendanceViewSet(viewsets.ModelViewSet):
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=False, methods=['post'])
    def check_in(self, request):
        employee = Employee.objects.get(user=request.user)
        today = timezone.now().date()
        attendance, created = Attendance.objects.get_or_create(
            employee=employee,
            date=today,
            defaults={'status': 'hadir', 'check_in': localtime(timezone.now()).time()}
        )
        if not created:
            return Response({'message': 'Sudah check in hari ini'}, status=status.HTTP_400_BAD_REQUEST)
        return Response(AttendanceSerializer(attendance).data)

    @action(detail=False, methods=['post'])
    def check_out(self, request):
        employee = Employee.objects.get(user=request.user)
        today = timezone.now().date()
        try:
            attendance = Attendance.objects.get(employee=employee, date=today)
            attendance.check_out = localtime(timezone.now()).time()
            attendance.save()
            return Response(AttendanceSerializer(attendance).data)
        except Attendance.DoesNotExist:
            return Response({'message': 'Belum check in hari ini'}, status=status.HTTP_400_BAD_REQUEST)
        
    def get_queryset(self):
        if self.request.user.is_staff:
            return Attendance.objects.filter(employee__user__is_staff=False)
        employee = Employee.objects.get(user=self.request.user)
        return Attendance.objects.filter(employee=employee)


class LeaveRequestViewSet(viewsets.ModelViewSet):
    queryset = LeaveRequest.objects.all()
    serializer_class = LeaveRequestSerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
        employee = Employee.objects.get(user=self.request.user)
        serializer.save(employee=employee)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        leave = self.get_object()
        leave.status = 'approved'
        leave.save()
        return Response(LeaveRequestSerializer(leave).data)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        leave = self.get_object()
        leave.status = 'rejected'
        leave.save()
        return Response(LeaveRequestSerializer(leave).data)
    
    def get_queryset(self):
        if self.request.user.is_staff:
            return LeaveRequest.objects.all()
        employee = Employee.objects.get(user=self.request.user)
        return LeaveRequest.objects.filter(employee=employee)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_attendance_excel(request):
    month = request.query_params.get('month')
    year = request.query_params.get('year')

    attendances = Attendance.objects.select_related('employee').all()
    if month:
        attendances = attendances.filter(date__month=month)
    if year:
        attendances = attendances.filter(date__year=year)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Rekap Absensi'

    headers = ['No', 'Nama Karyawan', 'ID Karyawan', 'Departemen', 'Tanggal', 'Check In', 'Check Out', 'Status']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for row, att in enumerate(attendances.order_by('date', 'employee__name'), 2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=att.employee.name)
        ws.cell(row=row, column=3, value=att.employee.employee_id)
        ws.cell(row=row, column=4, value=att.employee.department)
        ws.cell(row=row, column=5, value=str(att.date))
        ws.cell(row=row, column=6, value=str(att.check_in) if att.check_in else '-')
        ws.cell(row=row, column=7, value=str(att.check_out) if att.check_out else '-')
        ws.cell(row=row, column=8, value=att.status)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=rekap_absensi_{year}_{month}.xlsx'
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def rekap_per_karyawan(request):
    month = request.query_params.get('month')
    year = request.query_params.get('year')

    employees = Employee.objects.filter(is_active=True, user__is_staff=False)
    result = []

    for emp in employees:
        attendances = Attendance.objects.filter(employee=emp)
        if month:
            attendances = attendances.filter(date__month=month)
        if year:
            attendances = attendances.filter(date__year=year)

        hadir = attendances.filter(status='hadir').count()
        alpha = attendances.filter(status='alpha').count()

        leaves = LeaveRequest.objects.filter(employee=emp, status='approved')
        izin, sakit = hitung_hari_leave(leaves, month, year)

        result.append({
            'employee_id': emp.employee_id,
            'name': emp.name,
            'department': emp.department,
            'position': emp.position,
            'hadir': hadir,
            'alpha': alpha,
            'izin': izin,
            'sakit': sakit,
            'total': hadir + alpha + izin + sakit,
        })

    return Response(result)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_rekap_excel(request):
    month = request.query_params.get('month')
    year = request.query_params.get('year')

    employees = Employee.objects.filter(is_active=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Rekap Per Karyawan'

    headers = ['No', 'ID Karyawan', 'Nama', 'Departemen', 'Jabatan', 'Hadir', 'Alpha', 'Izin', 'Sakit', 'Total Hari']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for row, emp in enumerate(employees, 2):
        attendances = Attendance.objects.filter(employee=emp)
        if month:
            attendances = attendances.filter(date__month=month)
        if year:
            attendances = attendances.filter(date__year=year)

        hadir = attendances.filter(status='hadir').count()
        alpha = attendances.filter(status='alpha').count()

        leaves = LeaveRequest.objects.filter(employee=emp, status='approved')
        izin, sakit = hitung_hari_leave(leaves, month, year)

        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=emp.employee_id)
        ws.cell(row=row, column=3, value=emp.name)
        ws.cell(row=row, column=4, value=emp.department)
        ws.cell(row=row, column=5, value=emp.position)
        ws.cell(row=row, column=6, value=hadir)
        ws.cell(row=row, column=7, value=alpha)
        ws.cell(row=row, column=8, value=izin)
        ws.cell(row=row, column=9, value=sakit)
        ws.cell(row=row, column=10, value=hadir + alpha + izin + sakit)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=rekap_karyawan_{year}_{month}.xlsx'
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_employees_excel(request):
    employees = Employee.objects.filter(is_active=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Data Karyawan'

    headers = ['No', 'ID Karyawan', 'Nama', 'Username', 'Departemen', 'Jabatan', 'No HP', 'Tanggal Bergabung']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for row, emp in enumerate(employees, 2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=emp.employee_id)
        ws.cell(row=row, column=3, value=emp.name)
        ws.cell(row=row, column=4, value=emp.user.username if emp.user else '-')
        ws.cell(row=row, column=5, value=emp.department)
        ws.cell(row=row, column=6, value=emp.position)
        ws.cell(row=row, column=7, value=emp.phone)
        ws.cell(row=row, column=8, value=str(emp.join_date))

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=data_karyawan.xlsx'
    wb.save(response)
    return response